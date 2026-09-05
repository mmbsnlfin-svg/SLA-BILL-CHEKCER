
# ============================================================
# BSNL SLA BILL CHECKER - LOGIC ONLY (FOR STREAMLIT CLOUD / WEB)
# Same logic as Desktop V4.2
# Changes (notes + month parse + relaying retention option + MTTR export cleanup + duration conversion fix):
# 1) Month parsing FIX (supports Timestamp / YYYY-MM-DD)
# 2) Accounts Note PDF: final approved format
# 3) Penalty Clause 14.1 PDF: includes RKM, Rate, Total service value + SES blank line
# 4) 1% Relaying Not Done can be treated as Retention (checkbox from UI)
# 5) MTTR_Fault_Report: remove blank columns + add "Route Missing in A"
# 6) NEW: Route wise Fault Count Details table under Availability section in penalty note
# 7) FIX: Fault duration conversion now supports Excel hh:mm / hh:mm:ss / decimal reliably using direct cell read
# 8) NEW: PDF output instead of TXT note files
# 9) NEW: MTTR 25% cap calculated after deducting Availability Penalty from bill value
# ============================================================

import os
import re
import math
import calendar
from datetime import datetime, date, time, timedelta

import pandas as pd
import numpy as np
import openpyxl

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable, Preformatted
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib.enums import TA_CENTER


# -----------------------------
# Helpers
# -----------------------------
def safe_float(x, default=np.nan):
    try:
        s = str(x).strip()
        if s == "":
            return default
        return float(s)
    except Exception:
        return default


def norm_route_name(s):
    if pd.isna(s):
        return ""
    t = str(s).strip().lower()
    t = t.replace("\u00a0", " ")
    t = t.replace("–", "-").replace("—", "-")
    t = re.sub(r"\s+", " ", t)
    t = re.sub(r"\s*-\s*", "-", t)
    return t.strip()


def sanitize_filename(s, max_len=60):
    s = "" if s is None else str(s)
    s = s.strip()
    s = re.sub(r"[\\/:*?\"<>|]", "_", s)
    s = re.sub(r"\s+", " ", s)
    s = s.strip(" .-_")
    if len(s) == 0:
        s = "Unknown"
    return s[:max_len]


def ensure_engine(path):
    ext = os.path.splitext(path.lower())[1]
    if ext == ".xls":
        return "xlrd"
    return None


def read_excel_any(path, header=0):
    engine = ensure_engine(path)
    if engine:
        return pd.read_excel(path, header=header, engine=engine)
    return pd.read_excel(path, header=header)


def parse_duration_to_hours(val, number_format=None):
    """
    Accept duration in any of these forms:
    - decimal hours (e.g. 5.75 or 0.95)
    - HH:MM
    - HH:MM:SS
    - "1 day, 1:36:00"
    - datetime.time / python time object
    - pandas Timedelta / numpy timedelta64

    Important rule:
    When Excel is read directly with openpyxl, real time/duration cells usually
    come as time/timedelta objects, while pasted values come as numeric cells.
    So numeric cells are treated as DECIMAL HOURS and are NOT multiplied by 24.
    """
    if val is None:
        return np.nan

    try:
        if pd.isna(val):
            return np.nan
    except Exception:
        pass

    if isinstance(val, (timedelta, pd.Timedelta)):
        return val.total_seconds() / 3600.0

    if isinstance(val, np.timedelta64):
        return pd.to_timedelta(val).total_seconds() / 3600.0

    if isinstance(val, (datetime, pd.Timestamp)):
        return val.hour + (val.minute / 60.0) + (val.second / 3600.0)

    if isinstance(val, time):
        return val.hour + (val.minute / 60.0) + (val.second / 3600.0)

    # Numeric Excel cells are assumed to already be decimal hours.
    if isinstance(val, (int, float, np.integer, np.floating)):
        try:
            num = float(val)
        except Exception:
            return np.nan
        return np.nan if np.isnan(num) else num

    s = str(val).strip()
    if s == "":
        return np.nan

    if re.match(r"^\d{1,4}:\d{2}(:\d{2})?$", s):
        parts = s.split(":")
        h = int(parts[0])
        m = int(parts[1])
        sec = int(parts[2]) if len(parts) == 3 else 0
        return h + (m / 60.0) + (sec / 3600.0)

    if re.match(r"^\d+\s+days?\s+\d{1,2}:\d{2}(:\d{2})?$", s.lower()):
        td = pd.to_timedelta(s)
        return td.total_seconds() / 3600.0

    try:
        return float(s)
    except Exception:
        return np.nan


def get_excel_duration_series(excel_path, duration_header):
    """
    Read the duration column directly from Excel so time-formatted cells
    (hh:mm or hh:mm:ss) are converted reliably even if pandas changes type.
    """
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False, read_only=False)
        ws = wb[wb.sheetnames[0]]

        duration_header_norm = str(duration_header).strip()
        col_idx = None
        for c in range(1, ws.max_column + 1):
            hdr = ws.cell(1, c).value
            if str(hdr).strip() == duration_header_norm:
                col_idx = c
                break

        if col_idx is None:
            wb.close()
            return None

        values = []
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            values.append(parse_duration_to_hours(cell.value, cell.number_format))

        wb.close()
        return values
    except Exception:
        return None


def uptime_deduction_pct(uptime_pct):
    if pd.isna(uptime_pct):
        return 0
    if uptime_pct >= 99:
        return 0
    if uptime_pct >= 98:
        return 10
    if uptime_pct >= 97:
        return 25
    if uptime_pct >= 96:
        return 50
    if uptime_pct >= 95:
        return 75
    return 100


def mttr_penalty_non_cumulative(duration_hours):
    if pd.isna(duration_hours) or duration_hours <= 0:
        return 0, "Invalid/Blank"

    h = int(math.ceil(float(duration_hours)))  # ROUND UP
    if h <= 4:
        return 0, "≤4"
    elif h <= 6:
        return 500, ">4–6"
    elif h <= 24:
        return int(500 + 100 * int(math.ceil(h - 6))), ">6–24"
    elif h <= 48:
        return 5000, ">24–48"
    else:
        extra_days = int(math.ceil((h - 48) / 24))
        return int(5000 + 500 * extra_days), ">48"


def pan_4th_digit_to_tds_rate(pan4):
    if pan4 is None:
        return None
    s = str(pan4).strip().upper()
    if s == "":
        return None
    if s in ["P", "H"]:
        return 0.01
    return 0.02


def pick_first_nonblank(series):
    for v in series:
        if pd.isna(v):
            continue
        s = str(v).strip()
        if s != "":
            return s
    return ""


def detect_fault_duration_column(df):
    cols = list(df.columns)
    for col in cols:
        s = str(col).lower()
        if "fault" in s and "duration" in s:
            return col
    if len(cols) >= 14:
        return cols[13]  # Column N
    raise ValueError("Fault Duration column not found.")


def fmt_money(x):
    try:
        return f"{float(x):,.2f}"
    except Exception:
        return "0.00"


def robust_yes(val) -> bool:
    if pd.isna(val):
        return False
    s = str(val).strip().upper()
    if s == "":
        return False
    if s in {"NO", "N", "0", "FALSE"}:
        return False
    if s in {"YES", "Y", "1", "TRUE"}:
        return True
    if "YES" in s:
        return True
    if "EXEMPT" in s:
        return True
    return False


def find_exemption_column(columns):
    cols = list(columns)
    for col in cols:
        col_norm = str(col).strip().lower()
        if ("exempt" in col_norm) or ("exemption" in col_norm):
            return col

    for col in cols:
        col_norm = str(col).strip().lower()
        if (("mttr" in col_norm and "penalty" in col_norm)
            or ("avbility" in col_norm)
            or ("availability" in col_norm)):
            if ("yes" in col_norm or "no" in col_norm or "y/n" in col_norm or "yes/no" in col_norm):
                return col

    return None


def parse_month_year_from_value(val):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None, None

    if isinstance(val, (pd.Timestamp, datetime, date)):
        return int(val.year), int(val.month)

    s = str(val).strip()
    if s == "":
        return None, None

    m_iso = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m_iso:
        return int(m_iso.group(1)), int(m_iso.group(2))

    m = re.search(r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)", s.lower())
    y = re.search(r"(20\d{2})", s)
    mon_map = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,"jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
    if m and y:
        return int(y.group(1)), mon_map[m.group(1)]

    return None, None


def create_pdf_note(output_path, title, text_body):
    """
    Create a professional A4 portrait PDF note.
    Key penalty totals are RED + BOLD.
    Retention lines are BLUE + BOLD.
    """
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "BSNLTitle",
        parent=styles["Heading1"],
        alignment=TA_CENTER,
        fontSize=15,
        leading=19,
        textColor=colors.HexColor("#0b2d6b"),
        spaceAfter=8,
    )

    normal_style = ParagraphStyle(
        "BSNLMonoNormal",
        parent=styles["Code"],
        fontName="Courier",
        fontSize=8.4,
        leading=11.2,
        leftIndent=0,
        rightIndent=0,
        textColor=colors.black,
    )

    red_style = ParagraphStyle(
        "BSNLMonoRed",
        parent=normal_style,
        fontName="Courier-Bold",
        textColor=colors.red,
    )

    blue_style = ParagraphStyle(
        "BSNLMonoBlue",
        parent=normal_style,
        fontName="Courier-Bold",
        textColor=colors.HexColor("#0000CC"),
    )

    footer_style = ParagraphStyle(
        "BSNLFooter",
        parent=styles["BodyText"],
        alignment=TA_CENTER,
        fontSize=8,
        leading=10,
        textColor=colors.grey,
    )

    red_phrases = (
        "Total Penalty (MTTR + Availability)",
        "Total Retention Pending Scrutiny",
        "Total Net Penalty Recoverable",
        "Total Manual Penalties/Recoveries",
    )

    blue_phrases = (
        "MTTR Retention Pending Scrutiny",
        "Availability Retention Pending Scrutiny",
        "Retention:",
        "Retention for 1% Re-laying",
    )

    story = [
        Paragraph(title, title_style),
        HRFlowable(width="100%", thickness=0.8, color=colors.HexColor("#0b2d6b")),
        Spacer(1, 10),
    ]

    for raw_line in str(text_body).splitlines():
        if raw_line.strip() == "":
            story.append(Spacer(1, 5))
            continue

        if any(p in raw_line for p in red_phrases):
            style = red_style
        elif any(p in raw_line for p in blue_phrases):
            style = blue_style
        else:
            style = normal_style

        story.append(Preformatted(raw_line, style, maxLineLength=108))

    story.extend([
        Spacer(1, 14),
        HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey),
        Spacer(1, 6),
        Paragraph("Generated by BSNL SLA Bill Checker", footer_style),
    ])

    doc.build(story)


# -----------------------------
# Core Processing
# -----------------------------
def process_sla(
    annex_a_path,
    annex_c_path,
    rate_per_km,
    save_dir,
    vendor_basic_value=None,
    pan4=None,
    field_unit_penalty=0.0,
    vendor_deducted_penalty=0.0,
    other_recovery=0.0,
    splice_loss_amt=0.0,
    supervisor_abs_amt=0.0,
    frt_abs_amt=0.0,
    petroller_abs_amt=0.0,
    relaying_not_done_amt=0.0,
    relaying_as_retention=False,
):
    # ---------- Read Format A ----------
    a = read_excel_any(annex_a_path, header=0)
    a.columns = [str(c).strip() for c in a.columns]

    required_a = [
        "FORMAT", "BA", "OA", "Month", "Sr.No.",
        "Transnet Route ID", "Working Route Name as per Transnet",
        "RKM", "Name of Maintenance Agency"
    ]
    missing_a = [c for c in required_a if c not in a.columns]
    if missing_a:
        raise ValueError(f"Format A missing columns: {missing_a}. Ensure headers are exactly as finalized in Row-1.")

    routes = a[required_a].copy()
    routes.rename(columns={
        "Sr.No.": "Sl_No",
        "Transnet Route ID": "Route_ID",
        "Working Route Name as per Transnet": "Route_Name",
        "RKM": "Route_KM",
        "Name of Maintenance Agency": "Vendor_Name"
    }, inplace=True)

    routes["Route_ID"] = routes["Route_ID"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
    routes["Route_Name_norm"] = routes["Route_Name"].apply(norm_route_name)
    routes["Route_KM"] = pd.to_numeric(routes["Route_KM"], errors="coerce").fillna(0.0)

    rate_per_km = float(rate_per_km)
    routes["SLA_Charges_Rs"] = (routes["Route_KM"] * rate_per_km).round(4)

    ba_name = pick_first_nonblank(routes["BA"].tolist())
    oa_name = pick_first_nonblank(routes["OA"].tolist())
    vendor_name = pick_first_nonblank(routes["Vendor_Name"].tolist())

    sla_month_raw = routes["Month"].iloc[0] if len(routes) else ""
    year, month = parse_month_year_from_value(sla_month_raw)
    month_source = "Format-A" if (year is not None and month is not None) else "Pending Format-C fallback"

    # IMPORTANT: Do not default immediately to current month here.
    # Some field Format-A files can have a shifted/non-date value (e.g. "RKM")
    # in the Month column. In that case the actual service/fault month from
    # Format-C is used after Format-C is read.

    days_in_month = None
    total_hours_month = None
    month_name = None

    month_display = f"{month_name}-{year}"
    month_display_short = f"{month_name[:3]}-{year}"

    vendor_tag = sanitize_filename(vendor_name)
    month_tag2 = sanitize_filename(month_display)

    # Maps
    name_to_id = routes.set_index("Route_Name_norm")["Route_ID"].to_dict()
    id_to_name = routes.set_index("Route_ID")["Route_Name"].to_dict()
    route_ids_in_a = set(routes["Route_ID"])
    route_names_in_a = set(routes["Route_Name_norm"])

    # ---------- Read Format C ----------
    c = read_excel_any(annex_c_path, header=0)
    c.columns = [str(col).strip() for col in c.columns]

    must_have_c = ["Transnet Route ID", "Working Route Name as per Transnet"]
    missing_c = [x for x in must_have_c if x not in c.columns]
    if missing_c:
        raise ValueError(f"Format C missing required columns: {missing_c}. Ensure headers are exactly as finalized in Row-1.")

    duration_col = detect_fault_duration_column(c)

    faults = c.copy()
    faults["Route_ID_raw"] = faults["Transnet Route ID"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
    faults["Route_Name_raw"] = faults["Working Route Name as per Transnet"].astype(str)
    faults["Route_Name_norm"] = faults["Route_Name_raw"].apply(norm_route_name)

    exempt_col = find_exemption_column(faults.columns)
    faults["Is_Exempt"] = faults[exempt_col].apply(robust_yes) if exempt_col else False
    duration_from_excel = get_excel_duration_series(annex_c_path, duration_col)
    if duration_from_excel is not None and len(duration_from_excel) == len(faults):
        faults["Duration_Hrs"] = duration_from_excel
    else:
        faults["Duration_Hrs"] = faults[duration_col].apply(parse_duration_to_hours)

    faults_valid = faults[(faults["Duration_Hrs"].notna()) & (faults["Duration_Hrs"] > 0)].copy()
    faults_invalid = faults[~((faults["Duration_Hrs"].notna()) & (faults["Duration_Hrs"] > 0))].copy()

    # ---------- Resolve actual SLA month / calendar days ----------
    # Tender Annexure-C2 uses "Total days in Month x 24 hrs".
    # Therefore July=31 days, September=30 days, February=28/29 days, etc.
    # If Format-A Month is invalid/misaligned, use the valid Month from Format-C.
    fault_month_col = "Month" if "Month" in faults_valid.columns else None
    faults_valid["Fault_Year"] = np.nan
    faults_valid["Fault_Month"] = np.nan
    if fault_month_col:
        parsed_ym = faults_valid[fault_month_col].apply(parse_month_year_from_value)
        faults_valid["Fault_Year"] = parsed_ym.apply(lambda x: x[0] if x else None)
        faults_valid["Fault_Month"] = parsed_ym.apply(lambda x: x[1] if x else None)

        valid_month_pairs = [
            (int(y), int(m))
            for y, m in zip(faults_valid["Fault_Year"], faults_valid["Fault_Month"])
            if pd.notna(y) and pd.notna(m) and 1 <= int(m) <= 12
        ]
        if (year is None or month is None) and valid_month_pairs:
            # Monthly SLA input should normally contain one month. If more than one
            # appears, use the most frequent valid month as the billing-month fallback.
            from collections import Counter
            (year, month), _ = Counter(valid_month_pairs).most_common(1)[0]
            month_source = "Format-C fault month (fallback because Format-A Month is invalid)"

    if year is None or month is None:
        now = datetime.now()
        year, month = now.year, now.month
        month_source = "System month fallback (no valid month found in Format-A/Format-C)"

    days_in_month = calendar.monthrange(int(year), int(month))[1]
    total_hours_month = float(days_in_month * 24)
    month_name = calendar.month_name[int(month)]

    # Mapping ID first then Name
    faults_valid["Route_ID_mapped_by_id"] = faults_valid["Route_ID_raw"].where(faults_valid["Route_ID_raw"].isin(route_ids_in_a))
    faults_valid["Route_ID_mapped_by_name"] = faults_valid["Route_Name_norm"].map(name_to_id)

    faults_valid["Route_ID_Final"] = (
        faults_valid["Route_ID_mapped_by_id"]
        .fillna(faults_valid["Route_ID_mapped_by_name"])
        .fillna(faults_valid["Route_ID_raw"])
    )

    faults_valid["Matched_By_Name"] = np.where(
        faults_valid["Route_ID_mapped_by_id"].isna() & faults_valid["Route_ID_mapped_by_name"].notna(),
        "YES", "NO"
    )

    faults_valid["Route_Name_Final"] = faults_valid["Route_ID_Final"].map(id_to_name).fillna(faults_valid["Route_Name_raw"])

    # ---------- MTTR penalty ----------
    penalties = faults_valid["Duration_Hrs"].apply(mttr_penalty_non_cumulative)
    faults_valid["MTTR_Penalty_Tender_Rs"] = [p[0] for p in penalties]
    faults_valid["MTTR_Slab"] = [p[1] for p in penalties]

    faults_valid["MTTR_Penalty_Exempted_Rs"] = np.where(faults_valid["Is_Exempt"], faults_valid["MTTR_Penalty_Tender_Rs"], 0.0)
    faults_valid["MTTR_Penalty_Net_Rs"] = faults_valid["MTTR_Penalty_Tender_Rs"] - faults_valid["MTTR_Penalty_Exempted_Rs"]

    SLAB_ORDER = ["≤4", ">4–6", ">6–24", ">24–48", ">48"]
    slab_summary = (
        faults_valid.groupby("MTTR_Slab", dropna=False)
        .agg(
            Count=("MTTR_Slab", "size"),
            Penalty_Gross=("MTTR_Penalty_Tender_Rs", "sum"),
            Penalty_Exempted=("MTTR_Penalty_Exempted_Rs", "sum"),
            Penalty_Net=("MTTR_Penalty_Net_Rs", "sum"),
        )
        .reset_index()
    )
    all_slabs = pd.DataFrame({"MTTR_Slab": SLAB_ORDER})
    slab_summary = all_slabs.merge(slab_summary, on="MTTR_Slab", how="left").fillna(0)

    for col in ["Penalty_Gross", "Penalty_Exempted", "Penalty_Net"]:
        slab_summary[col] = slab_summary[col].round(2)

    total_faults_count = int(slab_summary["Count"].sum())
    mttr_gross = float(slab_summary["Penalty_Gross"].sum())
    mttr_exempt = float(slab_summary["Penalty_Exempted"].sum())
    mttr_net = float(slab_summary["Penalty_Net"].sum())

    # ---------- Availability ----------
    downtime_all = faults_valid.groupby("Route_ID_Final")["Duration_Hrs"].sum().reset_index()
    downtime_all.rename(columns={"Route_ID_Final": "Route_ID", "Duration_Hrs": "Downtime_Hrs_Total"}, inplace=True)

    faults_net = faults_valid[~faults_valid["Is_Exempt"]].copy()
    downtime_net = faults_net.groupby("Route_ID_Final")["Duration_Hrs"].sum().reset_index()
    downtime_net.rename(columns={"Route_ID_Final": "Route_ID", "Duration_Hrs": "Downtime_Hrs_Net"}, inplace=True)

    avail = routes.merge(downtime_all, on="Route_ID", how="left").merge(downtime_net, on="Route_ID", how="left")
    avail["Downtime_Hrs_Total"] = avail["Downtime_Hrs_Total"].fillna(0.0)
    avail["Downtime_Hrs_Net"] = avail["Downtime_Hrs_Net"].fillna(0.0)
    avail["Downtime_Exempted_Hrs"] = (avail["Downtime_Hrs_Total"] - avail["Downtime_Hrs_Net"]).round(4)

    # Route-wise month denominator. Prefer the actual valid fault month from Format-C.
    # This prevents a July route from being calculated on a 30-day/720-hour base.
    route_month_map = {}
    if "Fault_Year" in faults_valid.columns and "Fault_Month" in faults_valid.columns:
        month_rows = faults_valid[
            faults_valid["Fault_Year"].notna() & faults_valid["Fault_Month"].notna()
        ][["Route_ID_Final", "Fault_Year", "Fault_Month"]].copy()
        if len(month_rows):
            month_rows["YM"] = list(zip(month_rows["Fault_Year"].astype(int), month_rows["Fault_Month"].astype(int)))
            for rid, grp in month_rows.groupby("Route_ID_Final"):
                vals = grp["YM"].tolist()
                if vals:
                    from collections import Counter
                    route_month_map[str(rid)] = Counter(vals).most_common(1)[0][0]

    def _route_calendar_values(route_id):
        ry, rm = route_month_map.get(str(route_id), (int(year), int(month)))
        rd = calendar.monthrange(int(ry), int(rm))[1]
        return int(ry), int(rm), int(rd), float(rd * 24)

    route_cal = avail["Route_ID"].apply(_route_calendar_values)
    avail["Calc_Year"] = route_cal.apply(lambda x: x[0])
    avail["Calc_Month"] = route_cal.apply(lambda x: x[1])
    avail["Days_In_Month"] = route_cal.apply(lambda x: x[2])
    avail["Total_Hours_Month"] = route_cal.apply(lambda x: x[3])
    avail["Month_Used_For_Uptime"] = avail.apply(
        lambda r: f"{calendar.month_name[int(r['Calc_Month'])]}-{int(r['Calc_Year'])}", axis=1
    )

    # Use exact decimal uptime for deduction slab determination; do not round first.
    avail["Uptime_pct_Gross"] = ((avail["Total_Hours_Month"] - avail["Downtime_Hrs_Total"]) / avail["Total_Hours_Month"]) * 100.0
    avail["Uptime_pct_Net"] = ((avail["Total_Hours_Month"] - avail["Downtime_Hrs_Net"]) / avail["Total_Hours_Month"]) * 100.0
    avail["Uptime_pct_Gross"] = avail["Uptime_pct_Gross"].clip(0, 100)
    avail["Uptime_pct_Net"] = avail["Uptime_pct_Net"].clip(0, 100)

    avail["Deduction_pct_Gross"] = avail["Uptime_pct_Gross"].apply(uptime_deduction_pct)
    avail["Deduction_pct_Net"] = avail["Uptime_pct_Net"].apply(uptime_deduction_pct)

    avail["Deduction_Rs_Gross"] = (avail["SLA_Charges_Rs"] * avail["Deduction_pct_Gross"] / 100.0).round(2)
    avail["Deduction_Rs_Net"] = (avail["SLA_Charges_Rs"] * avail["Deduction_pct_Net"] / 100.0).round(2)

    avail["Availability_Deduction_Exempted_Rs"] = (avail["Deduction_Rs_Gross"] - avail["Deduction_Rs_Net"]).round(2)
    avail["Availability_Deduction_Net_Rs"] = avail["Deduction_Rs_Net"]

    availability_penalty_gross = round(float(avail["Deduction_Rs_Gross"].sum()), 2)
    availability_penalty_exempt = round(float(avail["Availability_Deduction_Exempted_Rs"].sum()), 2)
    availability_penalty_net = round(float(avail["Availability_Deduction_Net_Rs"].sum()), 2)

    # ---------- Missing routes (for notes) ----------
    missing_mask = (~faults_valid["Route_ID_Final"].isin(route_ids_in_a)) & (~faults_valid["Route_Name_norm"].isin(route_names_in_a))
    missing_rows = faults_valid[missing_mask].copy()
    missing_group = pd.DataFrame()
    if len(missing_rows) > 0:
        missing_group = (
            missing_rows.groupby(["Route_ID_raw", "Route_Name_raw"], dropna=False)["Duration_Hrs"]
            .sum()
            .reset_index()
            .rename(columns={"Route_ID_raw": "Route_ID_in_C", "Route_Name_raw": "Route_Name_in_C", "Duration_Hrs": "Total_Downtime_Hrs"})
            .sort_values("Total_Downtime_Hrs", ascending=False)
        )

    missing_lines = []
    if len(missing_group) == 0:
        missing_lines.append("Routes in Format-C but not found in Format-A: NIL")
    else:
        missing_lines.append("Routes in Format-C but not found in Format-A (even after ID + Name matching):")
        for _, r in missing_group.iterrows():
            missing_lines.append(f" - {r['Route_ID_in_C']} | {r['Route_Name_in_C']} | Downtime: {r['Total_Downtime_Hrs']:.2f} hrs")

    # ---------- MTTR 25% cap + retention pending scrutiny ----------
    # Tender logic:
    # - Availability penalty is outside the 25% MTTR cap.
    # - Cap base uses Availability Penalty BEFORE exemption.
    # - Gross MTTR is capped first, irrespective of exemption marking.
    # - Non-exempt MTTR is recoverable now.
    # - Remaining capped MTTR exposure attributable to exempt-marked faults
    #   is retained pending scrutiny.

    total_basic_sla = round(float(routes["SLA_Charges_Rs"].sum()), 2)

    mttr_cap_base_after_availability = round(
        total_basic_sla - availability_penalty_gross, 2
    )
    if mttr_cap_base_after_availability < 0:
        mttr_cap_base_after_availability = 0.0

    mttr_cap_25pct = round(mttr_cap_base_after_availability * 0.25, 2)

    mttr_penalty_after_cap = round(min(mttr_gross, mttr_cap_25pct), 2)
    mttr_cap_applied = "YES" if mttr_gross > mttr_cap_25pct else "NO"

    mttr_penalty_recoverable = round(
        min(mttr_net, mttr_penalty_after_cap), 2
    )

    mttr_retention_pending = round(
        min(
            mttr_exempt,
            max(mttr_penalty_after_cap - mttr_penalty_recoverable, 0.0)
        ),
        2
    )

    availability_penalty_recoverable = round(availability_penalty_net, 2)
    availability_retention_pending = round(availability_penalty_exempt, 2)

    total_penalty_before_mttr_cap = round(
        mttr_gross + availability_penalty_gross, 2
    )

    system_sla_penalty_gross = round(
        mttr_penalty_after_cap + availability_penalty_gross, 2
    )

    system_sla_retention_total = round(
        mttr_retention_pending + availability_retention_pending, 2
    )

    system_sla_penalty_net = round(
        mttr_penalty_recoverable + availability_penalty_recoverable, 2
    )

    system_sla_exemption_total = system_sla_retention_total
    mttr_net_after_cap = mttr_penalty_recoverable

    field_unit_penalty = float(field_unit_penalty or 0.0)
    higher_of_penalty = max(system_sla_penalty_net, field_unit_penalty)

    vendor_deducted_penalty = float(vendor_deducted_penalty or 0.0)
    sla_recovery_after_vendor = round(
        max(higher_of_penalty - vendor_deducted_penalty, 0.0), 2
    )

    # ---------- Accounts computations ----------
    total_rkm = round(float(routes["Route_KM"].sum()), 2)
    system_basic = round(total_basic_sla, 2)

    if vendor_basic_value is None or (isinstance(vendor_basic_value, float) and np.isnan(vendor_basic_value)):
        vendor_basic_used = system_basic
        vendor_basic_source = "System (RKM×Rate) (Vendor basic not provided)"
    else:
        vendor_basic_used = float(vendor_basic_value)
        vendor_basic_source = "Vendor Invoice"

    gst_on_vendor = round(vendor_basic_used * 0.18, 2)
    total_invoice_vendor = round(vendor_basic_used + gst_on_vendor, 2)

    gst_tds = round(vendor_basic_used * 0.02, 0)
    tds_rate = pan_4th_digit_to_tds_rate(pan4)
    if tds_rate is None:
        it_tds = 0.0
        it_tds_rate_text = "PAN 4th digit not provided (IT-TDS not computed by tool)"
    else:
        it_tds = round(vendor_basic_used * tds_rate, 0)
        it_tds_rate_text = f"{int(tds_rate*100)}% based on PAN 4th digit '{str(pan4).strip().upper()}'"

    total_statutory = round(gst_tds + it_tds, 0)
    net_before_penalty_and_vipa = round(total_invoice_vendor - total_statutory, 2)

    vipa_base = 1000.0 if vendor_basic_used > 500000 else 500.0
    vipa_gst = round(vipa_base * 0.18, 2)
    vipa_total = round(vipa_base + vipa_gst, 2)

    net_before_sla = round(total_invoice_vendor - gst_tds - it_tds - vipa_total, 2)

    # ---------- Manual inputs ----------
    splice_loss_amt = float(splice_loss_amt or 0.0)
    supervisor_abs_amt = float(supervisor_abs_amt or 0.0)
    frt_abs_amt = float(frt_abs_amt or 0.0)
    petroller_abs_amt = float(petroller_abs_amt or 0.0)
    relaying_not_done_amt = float(relaying_not_done_amt or 0.0)
    other_recovery = float(other_recovery or 0.0)

    relaying_penalty_amt = 0.0 if relaying_as_retention else relaying_not_done_amt
    relaying_retention_amt = relaying_not_done_amt if relaying_as_retention else 0.0

    total_penalty_clause14 = round(
        splice_loss_amt + system_sla_penalty_net +
        supervisor_abs_amt + frt_abs_amt + petroller_abs_amt + relaying_penalty_amt, 2
    )

    manual_penalties_accounts = round(
        splice_loss_amt + supervisor_abs_amt + frt_abs_amt + petroller_abs_amt + relaying_penalty_amt + other_recovery, 2
    )

    sla_retention_pending_total = round(system_sla_retention_total, 2)

    total_deductions_accounts = round(
        sla_recovery_after_vendor
        + sla_retention_pending_total
        + manual_penalties_accounts
        + relaying_retention_amt,
        2
    )
    net_payable_after_all = round(net_before_sla - total_deductions_accounts, 2)

    header_info = f"""BA: {ba_name}
OA: {oa_name}
SLA Month: {month_display}
Name of Maintenance Agency: {vendor_name}
"""

    # ============================================================
    # Accounts Note TXT (unchanged logic)
    # ============================================================
    relaying_line_accounts = (
        f"   Retention for 1% Re-laying work not done @200000/KM          = Rs. {fmt_money(relaying_retention_amt)}"
        if relaying_as_retention else
        f"   Penalty for 1% Re-laying work not done @200000/KM             = Rs. {fmt_money(relaying_penalty_amt)}"
    )

    retention_block_accounts = ""
    if relaying_as_retention and relaying_retention_amt > 0:
        retention_block_accounts = f"""
7) Retention:
{relaying_line_accounts}

"""

    accounts_note = f"""OFFICE NOTE

Subject: Verification of SLA Bill for {month_display}

BA: {ba_name}
OA: {oa_name}
SLA Month: {month_display}
Name of Maintenance Agency: {vendor_name}

1) Total route length as per Format-A:
   Total RKM                                                       = {total_rkm:.2f}

2) Rate as per Agreement / GeM:
   Rate                                                            = Rs. {rate_per_km:,.2f} per KM per month

   RKM x Rate ({total_rkm:.2f} x {rate_per_km:,.2f})               = Rs. {fmt_money(system_basic)}

3) Invoice Value (Accounts purpose):
   Basic Value (Source: {vendor_basic_source})                     = Rs. {fmt_money(vendor_basic_used)}
   Add: GST @18%                                                   = Rs. {fmt_money(gst_on_vendor)}
   -------------------------------------------------------------------------------
   Total Invoice Value                                             = Rs. {fmt_money(total_invoice_vendor)}

4) Statutory / Standard Deductions (rounded):
   GST TDS @2% on Basic                                            = Rs. {fmt_money(gst_tds)}
   TDS u/s 194C on Basic                                           = Rs. {fmt_money(it_tds)}  [{it_tds_rate_text}]
   -------------------------------------------------------------------------------
   Total Statutory Deduction (GST TDS + 194C TDS)                  = Rs. {fmt_money(total_statutory)}

   Net payable before penalty and VIPA (Invoice - Statutory)       = Rs. {fmt_money(net_before_penalty_and_vipa)}

   VIPA Charges (Auto)                                             = Rs. {fmt_money(vipa_base)}
   VIPA GST @18%                                                   = Rs. {fmt_money(vipa_gst)}
   VIPA Total                                                      = Rs. {fmt_money(vipa_total)}

A) Net payable to vendor (Before Penalty/Retention)                = Rs. {fmt_money(net_before_sla)}

5) SLA Penalty / Retention as per Tender and Competent Authority decision:

   MTTR PENALTY
   MTTR Penalty Before Exemption                                   = Rs. {fmt_money(mttr_gross)}
   MTTR Cap Base (Basic SLA - Availability Penalty Before Exemption)= Rs. {fmt_money(mttr_cap_base_after_availability)}
   Maximum MTTR Penalty as per 25% Rule                            = Rs. {fmt_money(mttr_cap_25pct)}
   MTTR Penalty after 25% Cap                                      = Rs. {fmt_money(mttr_penalty_after_cap)}
   MTTR Amount relating to faults marked Exempted                  = Rs. {fmt_money(mttr_exempt)}
   Net MTTR Penalty Recoverable                                    = Rs. {fmt_money(mttr_penalty_recoverable)}
   MTTR Retention Pending Scrutiny                                 = Rs. {fmt_money(mttr_retention_pending)}
   Cap Applied                                                     = {mttr_cap_applied}

   AVAILABILITY PENALTY
   Availability Penalty Before Exemption                           = Rs. {fmt_money(availability_penalty_gross)}
   Availability Amount relating to Exempted Faults                 = Rs. {fmt_money(availability_penalty_exempt)}
   Net Availability Penalty Recoverable                            = Rs. {fmt_money(availability_penalty_recoverable)}
   Availability Retention Pending Scrutiny                         = Rs. {fmt_money(availability_retention_pending)}

   -------------------------------------------------------------------------------
   Total Penalty (MTTR + Availability)                             = Rs. {fmt_money(system_sla_penalty_gross)}
   Total Retention Pending Scrutiny                                = Rs. {fmt_money(system_sla_retention_total)}
   Total Net Penalty Recoverable                                   = Rs. {fmt_money(system_sla_penalty_net)}
   -------------------------------------------------------------------------------

   Penalty as per Field Unit / SES (Info)                          = Rs. {fmt_money(field_unit_penalty)}
   Adopted Penalty (Higher of Net SLA / Field Unit)                = Rs. {fmt_money(higher_of_penalty)}

   Vendor already deducted SLA penalty (if any)                    = Rs. {fmt_money(vendor_deducted_penalty)}
   Net SLA recovery after vendor deduction                         = Rs. {fmt_money(sla_recovery_after_vendor)}

   Note: Retention against faults marked as exempted is pending scrutiny.
   If exemption is approved by Competent Authority, applicable retention may be released.
   If exemption is disallowed, applicable retained amount may be adjusted as penalty/LD.

6) Manual Penalties / Recoveries:
   Splice loss per fiber                                           = Rs. {fmt_money(splice_loss_amt)}
   Absence of Supervisor @1500/day                                 = Rs. {fmt_money(supervisor_abs_amt)}
   Absence of FRT @5000/day                                        = Rs. {fmt_money(frt_abs_amt)}
   Absence of Petroller @500/day                                   = Rs. {fmt_money(petroller_abs_amt)}
{relaying_line_accounts}
   Any other recovery                                              = Rs. {fmt_money(other_recovery)}
   -------------------------------------------------------------------------------
   Total Manual Penalties/Recoveries                               = Rs. {fmt_money(manual_penalties_accounts)}
{retention_block_accounts}
7) SLA Retention Pending Scrutiny                                  = Rs. {fmt_money(sla_retention_pending_total)}

B) Total Deductions (Penalty + SLA Retention + Manual + Other Retention)
                                                                    = Rs. {fmt_money(total_deductions_accounts)}

Net Payable to Vendor (A - B)                                      = Rs. {fmt_money(net_payable_after_all)}

8) Route mapping remarks:
{chr(10).join(missing_lines)}

The following documents have been verified and uploaded with MIRO transaction:
1. Invoice along with supporting documents.
2. All other documents attached by user unit in SES (pl go through it)

Bill put up for approval please.

Submitted for approval.
"""

    # ============================================================
    # Clause 14.1 Technical Note (Penalty Note)
    # ============================================================
    slab_map = {
        "≤4": "Upto 4 Hrs",
        ">4–6": "Between 4 Hrs to 6 Hrs",
        ">6–24": "Between 6 Hrs  to 24 Hrs",
        ">24–48": "Between 24 hrs to 48 Hrs",
        ">48": "Beyond 48 Hrs."
    }

    mttr_header = f"{'Slab':<30} {'Count':>7} {'Penalty':>14} {'Exempted':>14} {'Net':>14}"
    mttr_sep = "-" * len(mttr_header)
    mttr_lines = [mttr_header, mttr_sep]
    for _, r in slab_summary.iterrows():
        name = slab_map.get(r["MTTR_Slab"], str(r["MTTR_Slab"]))
        mttr_lines.append(
            f"{name:<30} "
            f"{int(r['Count']):>7} "
            f"{fmt_money(r['Penalty_Gross']):>14} "
            f"{fmt_money(r['Penalty_Exempted']):>14} "
            f"{fmt_money(r['Penalty_Net']):>14}"
        )
    mttr_lines.append(mttr_sep)
    mttr_lines.append(
        f"{'Total':<30} "
        f"{total_faults_count:>7} "
        f"{fmt_money(mttr_gross):>14} "
        f"{fmt_money(mttr_exempt):>14} "
        f"{fmt_money(mttr_net):>14}"
    )

    avail_view = avail.copy()
    avail_view["Route_ID"] = avail_view["Route_ID"].astype(str)
    avail_view["Route_Name"] = avail_view["Route_Name"].astype(str)

    avail_focus = avail_view[(avail_view["Availability_Deduction_Net_Rs"] > 0) | (avail_view["Downtime_Hrs_Total"] > 0)].copy()
    avail_focus = avail_focus.sort_values(["Availability_Deduction_Net_Rs", "Downtime_Hrs_Net"], ascending=[False, False])

    av_header = f"{'Route ID':<18} {'Route Name':<55} {'Uptime%':>8} {'Ded%':>6} {'Penalty(Net)':>14}"
    av_sep = "-" * len(av_header)
    av_lines = [av_header, av_sep]

    if len(avail_focus) == 0:
        av_lines.append("No route has downtime/penalty for this month.")
    else:
        for _, rr in avail_focus.iterrows():
            rid = str(rr["Route_ID"])[:18]
            rname = str(rr["Route_Name"])[:55]
            av_lines.append(
                f"{rid:<18} {rname:<55} "
                f"{rr['Uptime_pct_Net']:>8.2f} "
                f"{int(rr['Deduction_pct_Net']):>6} "
                f"{fmt_money(rr['Availability_Deduction_Net_Rs']):>14}"
            )

    # ✅ NEW: Route wise Fault Count Details table
    fault_count_df = (
        faults_valid.groupby("Route_ID_Final")
        .size()
        .reset_index(name="Total_Fault_Count")
        .rename(columns={"Route_ID_Final": "Route_ID"})
    )
    fault_count_df = fault_count_df.merge(
        avail_view[["Route_ID", "Route_Name", "Uptime_pct_Net"]],
        on="Route_ID",
        how="left"
    )
    fault_count_df["Route_Name"] = fault_count_df["Route_Name"].fillna("")
    fault_count_df["Uptime_pct_Net"] = pd.to_numeric(fault_count_df["Uptime_pct_Net"], errors="coerce").fillna(0.0)
    fault_count_df["Total_Fault_Count"] = pd.to_numeric(fault_count_df["Total_Fault_Count"], errors="coerce").fillna(0).astype(int)
    fault_count_df = fault_count_df.sort_values(["Total_Fault_Count", "Uptime_pct_Net"], ascending=[False, True])

    fc_header = f"{'Route ID':<18} {'Route Name':<55} {'Total Fault count':>18} {'Uptime%':>10}"
    fc_sep = "-" * len(fc_header)
    fc_lines = [fc_header, fc_sep]

    total_faults_all_routes = int(fault_count_df["Total_Fault_Count"].sum())
    for _, rr in fault_count_df.iterrows():
        rid = str(rr["Route_ID"])[:18]
        rname = str(rr["Route_Name"])[:55]
        cnt = int(rr["Total_Fault_Count"])
        up = float(rr["Uptime_pct_Net"])
        fc_lines.append(f"{rid:<18} {rname:<55} {cnt:>18} {up:>10.2f}")

    fc_lines.append(fc_sep)
    fc_lines.append(f"{'Total =':<75} {total_faults_all_routes}")

    relaying_line_penalty_note = (
        f"7. Retention for 1% Re-laying ofc Work not done @ 200000 Per KM Rs.  : Rs. {fmt_money(relaying_retention_amt)}"
        if relaying_as_retention else
        f"7. Penalty for 1% Re-laying ofc Work not done @ 200000 Per KM Rs.    : Rs. {fmt_money(relaying_penalty_amt)}"
    )

    total_deduction_penalty_note = round(
        total_penalty_clause14 + sla_retention_pending_total + relaying_retention_amt, 2
    )

    technical_note_clause14 = f"""SLA PENALTIES CALCULATION AS PER TENDER CLAUSE 14.1

{header_info}

Total of RKM as per Annexure A                                   = {total_rkm:.2f}

Rate as per Tender Rs.                                           = {rate_per_km:,.2f} Per KM Monthly.

Total Value of  service Rs. (RKM*RATE)                           = Rs. {fmt_money(system_basic)}

Penalty Details given below:-

1. SPLICE LOSS PER FIBER Rs.                                     : Rs. {fmt_money(splice_loss_amt)}

2. MTTR FAULTS Penalty Details                                   : (Code generated from Format-C)

   MTTR Cap Base after Availability Penalty BEFORE Exemption Rs.   : Rs. {fmt_money(mttr_cap_base_after_availability)}
   Max 25% MTTR Penalty Rs.                                        : Rs. {fmt_money(mttr_cap_25pct)}
   MTTR Penalty after 25% Cap                                      : Rs. {fmt_money(mttr_penalty_after_cap)}

   Slab wise faults and penalty:
{chr(10).join(mttr_lines)}

   Net MTTR Penalty Recoverable                                  : Rs. {fmt_money(mttr_penalty_recoverable)}
   MTTR Retention Pending Scrutiny                                : Rs. {fmt_money(mttr_retention_pending)}
   Cap Applied                                                    : {mttr_cap_applied}

3. Availability Penalty Before Exemption                         : Rs. {fmt_money(availability_penalty_gross)}
   Availability Amount relating to Exempted Faults                 : Rs. {fmt_money(availability_penalty_exempt)}
   Net Availability Penalty Recoverable                            : Rs. {fmt_money(availability_penalty_recoverable)}
   Availability Retention Pending Scrutiny                         : Rs. {fmt_money(availability_retention_pending)}

   Availability details (Net):
{chr(10).join(av_lines)}

   Route wise Fault Count Details
{chr(10).join(fc_lines)}

   SLA Penalty Summary
   Total Penalty (MTTR + Availability)                            : Rs. {fmt_money(system_sla_penalty_gross)}
   Total Retention Pending Scrutiny                               : Rs. {fmt_money(system_sla_retention_total)}
   Total Net Penalty Recoverable                                  : Rs. {fmt_money(system_sla_penalty_net)}

4. Absense of Supervisor @ 1500 per day Rs.                       : Rs. {fmt_money(supervisor_abs_amt)}
5. Absence of FRT @ 5000 Per day Rs.                              : Rs. {fmt_money(frt_abs_amt)}
6. Absence of Petroller @ 500 Per day Rs.                         : Rs. {fmt_money(petroller_abs_amt)}
{relaying_line_penalty_note}

Total Penalty (1+2+3+4+5+6+7 as applicable) Rs.                                   : Rs. {fmt_money(total_penalty_clause14)}
Total SLA Retention Pending Scrutiny Rs.                         : Rs. {fmt_money(sla_retention_pending_total)}
Total Deduction (Penalty + SLA Retention + Other Retention) Rs.    : Rs. {fmt_money(total_deduction_penalty_note)}

The supporting documents listed below have been verified
and uploaded under SAP Service Entry Sheet No: ______________

1. Excel file of Annexure A
2. Excel file of Annexure C

Submitted for approval please.
"""

    # ---------- Output files ----------
    os.makedirs(save_dir, exist_ok=True)
    out_xlsx = os.path.join(save_dir, f"SLA_Output_{vendor_tag}_{month_tag2}.xlsx")
    out_accounts_pdf = os.path.join(save_dir, f"SAP_Accounts_Note_{vendor_tag}_{month_tag2}.pdf")
    out_tech_pdf = os.path.join(save_dir, f"Penalty_Clause14_1_{vendor_tag}_{month_tag2}.pdf")

    summary = pd.DataFrame([
        ["BA", ba_name],
        ["OA", oa_name],
        ["Vendor", vendor_name],
        ["SLA Month (Format-A raw)", str(sla_month_raw)],
        ["Month source used", month_source],
        ["Default month used for routes without valid fault month", f"{month_name} {year} ({days_in_month} days)"],
        ["Uptime denominator rule", "Actual calendar days of route/fault month x 24 hrs"],
        ["Total RKM", total_rkm],
        ["Rate per KM", rate_per_km],
        ["Total Basic SLA (Σ RKM×Rate)", round(float(routes["SLA_Charges_Rs"].sum()), 2)],
        ["MTTR cap base after availability penalty BEFORE exemption", mttr_cap_base_after_availability],
        ["MTTR cap 25%", mttr_cap_25pct],
        ["MTTR cap applied", mttr_cap_applied],
        ["MTTR gross before exemption", mttr_gross],
        ["MTTR amount relating to exempt-marked faults", mttr_exempt],
        ["MTTR penalty after 25% cap", mttr_penalty_after_cap],
        ["MTTR penalty recoverable", mttr_penalty_recoverable],
        ["MTTR retention pending scrutiny", mttr_retention_pending],
        ["Availability penalty gross", availability_penalty_gross],
        ["Availability amount relating to exempt-marked faults", availability_penalty_exempt],
        ["Availability penalty recoverable", availability_penalty_recoverable],
        ["Availability retention pending scrutiny", availability_retention_pending],
        ["System SLA penalty gross", system_sla_penalty_gross],
        ["System SLA retention pending scrutiny", system_sla_retention_total],
        ["System SLA penalty recoverable", system_sla_penalty_net],
        ["Field unit penalty (info)", field_unit_penalty],
        ["Higher-of adopted penalty", higher_of_penalty],
        ["Vendor already deducted SLA penalty", vendor_deducted_penalty],
        ["Net SLA recovery after vendor deduction", sla_recovery_after_vendor],
        ["Clause 14.1 penalty total (excluding retention)", total_penalty_clause14],
        ["Relaying treated as retention", "YES" if relaying_as_retention else "NO"],
        ["Relaying retention amount", relaying_retention_amt],
        ["Valid faults count", len(faults_valid)],
        ["Exempt faults count", int(faults_valid["Is_Exempt"].sum())],
        ["Invalid duration rows", len(faults_invalid)],
        ["Duration column used", str(duration_col)],
        ["Exemption column used", str(exempt_col) if exempt_col else "None"],
    ], columns=["Item", "Value"])

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        avail.sort_values(["Route_ID"]).to_excel(writer, sheet_name="Availability_Report", index=False)

        # ✅ MTTR_Fault_Report CLEAN EXPORT + NEW COLUMN
        mttr_export = faults_valid.copy()
        mttr_export["Route Missing in A"] = np.where(
            mttr_export["Route_ID_Final"].isin(route_ids_in_a),
            "NO",
            "YES"
        )

        drop_cols = [
            "Route_ID_mapped_by_id",
            "Route_ID_mapped_by_name",
            "Route_ID_Final",
            "Route_Name_norm",
        ]
        for col in drop_cols:
            if col in mttr_export.columns:
                mttr_export.drop(columns=col, inplace=True)

        mttr_export = mttr_export.dropna(axis=1, how="all")

        cols = list(mttr_export.columns)
        if "Route Missing in A" in cols:
            cols.remove("Route Missing in A")
            cols.append("Route Missing in A")
            mttr_export = mttr_export[cols]

        mttr_export.to_excel(writer, sheet_name="MTTR_Fault_Report", index=False)

        slab_summary.to_excel(writer, sheet_name="MTTR_Slab_Summary", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)

        if len(faults_invalid) > 0:
            faults_invalid.to_excel(writer, sheet_name="Invalid_Fault_Rows", index=False)

    create_pdf_note(out_accounts_pdf, "OFFICE NOTE", accounts_note)
    create_pdf_note(out_tech_pdf, "SLA PENALTY NOTE", technical_note_clause14)

    return out_xlsx, out_accounts_pdf, out_tech_pdf
